from flask import Flask, flash, redirect, url_for, render_template, request
from werkzeug import secure_filename
import os
import openpyxl
from openpyxl import load_workbook
import MySQLdb

encoding = "utf-8" 
on_error = "replace"

DATABASE="ManageBacdb"
HOST="localhost"
USERNAME="dodo"
PASSWORD="Dodo-131459"

app = Flask(__name__)
app.secret_key = 'some_secret'

app.config['UPLOAD_FOLDER'] = '/home/richardo/FlaskApp/uploads'
app.config['MAX_CONTENT_PATH'] = 10000000

def OpenDb():
	global db, cursor
	db = MySQLdb.connect(HOST,USERNAME,PASSWORD,DATABASE,charset='utf8',use_unicode=True)
	cursor = db.cursor()

def CloseDb():
	global db, cursor
	cursor.close()
	db.close()
	
@app.route("/")
def main():
    return render_template('FormImport.html')

@app.route("/upload", methods=['GET','POST'])
def upload():
	if request.method == 'POST':
		f = request.files['file']
		StartYear = request.form['StartYear']
		Grade = request.form['Grade']
		FileName = app.config['UPLOAD_FOLDER'] + '/' + secure_filename(f.filename)
		try:
			f.save(FileName)
		except:
			flash('Upload Failed')
			return redirect(url_for('main'))
			#return render_template('upload_gagal.html', FileName=secure_filename(f.filename))
		
		if ImportFile(StartYear, Grade, FileName) == 'exist':
			flash('File Already Uploaded')
			return redirect(url_for('main'))
			#return render_template('import_exist.html', FileName=secure_filename(f.filename))
		else:
			flash('Import Success')
			return redirect(url_for('main'))
			#return render_template('import_sukses.html', FileName=secure_filename(f.filename))
        
	
def ImportFile(StartYear, Grade, FileName):
	wb = load_workbook(FileName)
	sheet = wb.get_sheet_by_name('Diploma by Student')
	row = 7
	JmlBaris = sheet.max_row + 1
	name = sheet['A5'].value
	text1 = sheet['A3'].value
	parse1 = text1.split(" ")
	parse2 = parse1[5].split("(")
	StartMonth = str(parse2[1]).strip(' \t\n\r')
	FileStartYear = str(parse1[6]).strip(' \t\n\r')
	parse3 = parse1[9].split(")")
	EndMonth = str(parse1[8]).strip(' \t\n\r')
	EndYear = str(parse3[0]).strip(' \t\n\r')
	
	if str(parse1[3]) == "Semester":
		semester = str(parse1[4]) + "B"
	else:
		semester = str(parse1[4]) + "A"
	
	TermId = StartYear+"GR"+Grade+semester
	
	ValidateQry = "SELECT TermId FROM TermGrades WHERE TermId ='"
	ValidateQry += TermId
	ValidateQry += "'"
	OpenDb()
	cursor.execute(ValidateQry)
	data = cursor.fetchone()
	CloseDb()
	
	if data == None :
			
		while row <= JmlBaris:
			if sheet['A' + str(row)].value == None or sheet['A' + str(row)].value == "":
				name = sheet['A' + str(row+1)].value
				row = row+3
				continue
			
			else:
				#name = sheet['A' + str(row-2)].value
				F = sheet['A' + str(row)].value
				G = sheet['B' + str(row)].value
				H = sheet['C' + str(row)].value
				I = sheet['D' + str(row)].value
				J = sheet['E' + str(row)].value
				#K = sheet['F' + str(row)].value --Grade Value
				L = sheet['G' + str(row)].value
				M = sheet['H' + str(row)].value
				O = sheet['I' + str(row)].value
				P = sheet['J' + str(row)].value
				Q = sheet['K' + str(row)].value
				R = sheet['L' + str(row)].value
				S = sheet['M' + str(row)].value
				T = sheet['N' + str(row)].value
				U = sheet['O' + str(row)].value
				V = sheet['P' + str(row)].value
				W = sheet['Q' + str(row)].value
				X = sheet['R' + str(row)].value
				Y = sheet['S' + str(row)].value
				comments = Y.replace("'", "")
				#sheet1.append([StartMonth,StartYear,EndMonth,EndYear,semester,F,G,H,I,J,K,L,M,str(name),O,P,Q,R,S,T,U,V,W,X,Y])
				try :
					OpenDb()
					insQry = "Insert into TermGrades VALUES ('"	
					insQry += TermId
					insQry += "','"
					insQry += StartYear
					insQry += "','"
					insQry += StartMonth
					insQry += "','"
					insQry += EndYear
					insQry += "','"
					insQry += EndMonth
					insQry += "','"
					insQry += semester
					insQry += "','"
					insQry += F
					insQry += "','"
					insQry += G
					insQry += "','"
					insQry += H
					insQry += "','"
					insQry += I
					insQry += "','"
					insQry += J
					insQry += "','"
					insQry += Grade
					insQry += "','"
					insQry += L
					insQry += "','"
					insQry += M
					insQry += "','"
					insQry += str(name)
					insQry += "','"
					insQry += str(O)
					insQry += "','"
					insQry += str(P)
					insQry += "','"
					insQry += Q
					insQry += "','"
					insQry += R
					insQry += "','"
					insQry += S
					insQry += "','"
					insQry += T
					insQry += "','"
					insQry += U
					insQry += "','"
					insQry += V
					insQry += "','"
					insQry += W
					insQry += "','"
					insQry += X
					insQry += "','"
					insQry += comments
					insQry += "')"
					cursor.execute(insQry)
					db.commit()
					CloseDb()
					
				
				except MySQLdb.Error as err:
					flash('Import Failed')
					return redirect(url_for('main'))
				
			row = row+1
		
	else :
		return 'exist'
		

@app.route("/search_form")
def search_form():
    return render_template('form_search.html')
		
@app.route("/search", methods=['GET','POST'])
def search():
	if request.method == 'POST':
		StudentId = request.form['StudentId']
		SelectQryName = "Select StudentName FROM TermGrades Where StudentID='"	
		SelectQryName += StudentId
		SelectQryName += "'"
		OpenDb()
		cursor.execute(SelectQryName)
		result = cursor.fetchone()
		if result == None or result == "":
			#messages='Student ID not found'
			flash('Student ID not found')
			return redirect(url_for('search_form'))
		else:
			StudentName = str(cursor.fetchone()[0])
			CloseDb()
		
		return ViewGrades(StudentId, StudentName)
		
		
def ViewGrades(StudentId, StudentName):
	TempDb(StudentId)
	InsertGrades(StudentId)
	Records = ViewRecords(StudentId)
	Years = ViewYears(StudentId)
	Year = []
	for x in range(len(Years)):
		year = str(Years[x])
		if str(year) == 'N/A' :
			Year.append('N/A')
		else:
			EndYear=(int(year) + 1)
			Year.append(str(year) + " - " + str(EndYear))
		
	return render_template('form_view.html', StudentId=StudentId, StudentName=StudentName, container=Records, year9=Year[0], year10=Year[1], year11=Year[2], year12=Year[3])

def TempDb(StudentId):
	OpenDb()
	cursor.execute("DELETE FROM TempView WHERE StudentId = '"+StudentId+"'")
	cursor.execute("DELETE FROM TempYear WHERE StudentId = '"+StudentId+"'")
	#cursor.execute("INSERT INTO TempView VALUES('A','A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A')")
	db.commit()
	CloseDb()
	grade = 9
	while grade <= 12:
		try :
			SelectQrySubject = "Select ClassId, Subject FROM TermGrades Where StudentID='"	
			SelectQrySubject += StudentId
			SelectQrySubject += "' AND GradeLevel = '"
			SelectQrySubject += str(grade)
			SelectQrySubject += "'"
			OpenDb()
			cursor.execute(SelectQrySubject)
			#container = cursor.fetchall()
			data = cursor.fetchone()
			if data == None or data == "":
				grade = grade + 1
				continue
			else:
				container = cursor.fetchall()
				for row in container:
					InsertQry = "INSERT IGNORE INTO TempView VALUES ('"
					InsertQry += str(row[0])
					InsertQry += "','" 
					InsertQry += str(row[1])
					InsertQry += "','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','"
					InsertQry += StudentId
					InsertQry += "') ON DUPLICATE KEY UPDATE Subject='"
					InsertQry += str(row[1])
					InsertQry += "'"
					cursor.execute(InsertQry)
					db.commit()
					#CloseDb()
					grade = grade + 1
								
		except MySQLdb.Error as err:
			return err
	
	CloseDb()
	return container
	
def InsertGrades(StudentId):
	OpenDb()
	cursor.execute("INSERT INTO TempYear VALUES('N/A','N/A','N/A','N/A','"+StudentId+"')")
	db.commit()
	cursor.execute("Select ClassId FROM TempView")
	ClassAll = cursor.fetchall()
	for ClassId in ClassAll:
		grade = 9
		while grade <= 12:
			SelectQry = "Select StartYear, GradeLevel, Semester, FinalGrade FROM TermGrades WHERE StudentID='"	
			SelectQry += StudentId
			SelectQry += "' AND ClassId='"
			SelectQry += str(ClassId[0])
			SelectQry += "' AND TermId LIKE '%GR"
			SelectQry += str(grade)
			SelectQry += "%'"
			cursor.execute(SelectQry)
			data = cursor.fetchall()
			if data == None:
				grade = grade + 1
				continue
				
			else:
				for row in data:
					#Year = str(grade) + str(data[0])
					Semester = row[2]
					UpdateYear = "UPDATE TempYear SET Year"
					UpdateYear += str(grade)
					UpdateYear += "= '"
					UpdateYear += str(row[0])
					UpdateYear += "'"
					cursor.execute(UpdateYear)
					db.commit()
					UpdateQry = "UPDATE TempView SET Grade"
					UpdateQry += str(grade)
					UpdateQry += str(Semester)
					UpdateQry += " = '"
					UpdateQry += str(row[3])
					UpdateQry += "' WHERE ClassId='"
					UpdateQry += str(ClassId[0])
					UpdateQry += "'"
					cursor.execute(UpdateQry)
					db.commit()
				grade = grade + 1
	CloseDb()
	
def ViewRecords(StudentId):
	OpenDb()
	cursor.execute("SELECT * FROM TempView WHERE StudentId = '"+StudentId+"'")
	Records = cursor.fetchall()
	CloseDb()
	return Records
	
def ViewYears(StudentId):
	OpenDb()
	cursor.execute("SELECT * FROM TempYear WHERE StudentId = '"+StudentId+"'")
	Years = cursor.fetchone()
	CloseDb()
	return Years
	
@app.route("/acceptValue" ,methods=['POST'])
def acceptValue():
	StudentId=request.form['StudentId']
	StudentName=request.form['StudentName']
	GradeLevel=request.form['GradeLevel']
	Grade = GradeLevel[2:4]
	Semester = GradeLevel[4:6]
	ClassId=request.form['ClassId']
	OpenDb()
	cursor.execute("SELECT Subject FROM TempView WHERE ClassId = '"+ClassId+"'")
	Subject=str(cursor.fetchone()[0])
	CloseDb()
	FinalGrade=request.form['FinalGrade']
	OpenDb()
	cursor.execute("SELECT * FROM TermGrades WHERE TermId Like '%"+GradeLevel+"'")
	TermId=str(cursor.fetchone()[0])
	CloseDb()
	if TermId == None or TermId == "":
		flash('Grade Record not found')
		return redirect(url_for('search'))
	else:
		return render_template('edit_form.html', TermId=TermId, StudentId=StudentId, StudentName=StudentName, GradeLevel=GradeLevel, Grade=Grade, Semester=Semester, ClassId=ClassId, Subject=Subject, FinalGrade=FinalGrade)

@app.route("/UpdateGrade" ,methods=['POST'])
def UpdateGrade():
	TermId=request.form['TermId']
	StudentId=request.form['StudentId']
	ClassId=request.form['ClassId']
	NewGrade=request.form['NewGrade']
	OpenDb()
	cursor.execute("Select StudentName FROM TermGrades Where StudentID='"+StudentId+"'")
	StudentName=str(cursor.fetchone()[0])
	CloseDb()
	OpenDb()
	UpdateGrade = "UPDATE TermGrades SET FinalGrade = '"
	UpdateGrade += str(NewGrade)
	UpdateGrade += "' WHERE TermId = '"
	UpdateGrade += str(TermId)
	UpdateGrade += "' AND ClassId = '"
	UpdateGrade += str(ClassId)
	UpdateGrade += "' AND StudentID = '"
	UpdateGrade += str(StudentId)
	UpdateGrade += "'"
	cursor.execute(UpdateGrade)
	db.commit()
	CloseDb()
	flash('Grade Record Update Success')
	return ViewGrades(StudentId, StudentName)
	
    
if __name__ == "__main__":
    app.run()
